"""
Prikat part of CisLink integration
"""
import sys
import time
import os
from ftplib import FTP
from typing import List

import ms
import openpyxl
# pip install openpyxl
import csv

from settings import CL_PRIKAT_NAME, CL_USER, CL_PASSWD, DIR, CL_XLSX_TEMPLATE, CL_XLSX_TEMPLATE_LIST, \
    CL_OUR_GLN, CL_ED_IZM_RAZM, CL_ED_IZM_VOLUME, CL_ED_IZM_WEIGHT, CL_CURRENCY, CL_ED_IZM, URL_FTP, TRY, \
    CL_PRICAT_FOLDER


def make_xlsx(res: List) -> str:
    wb = openpyxl.load_workbook(DIR + CL_XLSX_TEMPLATE)
    sheet = wb[CL_XLSX_TEMPLATE_LIST]
    i = 1
    for elem in res:
        sheet.cell(row=i + 1, column=1).value = CL_OUR_GLN
        sheet.cell(row=i + 1, column=2).value = i
        sheet.cell(row=i + 1, column=3).value = elem['barcode']
        sheet.cell(row=i + 1, column=5).value = elem['code']
        sheet.cell(row=i + 1, column=6).value = elem['name']
        sheet.cell(row=i + 1, column=7).value = 1
        sheet.cell(row=i + 1, column=8).value = CL_ED_IZM
        sheet.cell(row=i + 1, column=9).value = elem['group']
        sheet.cell(row=i + 1, column=10).value = elem['brand']
        sheet.cell(row=i + 1, column=14).value = elem['glubina']
        sheet.cell(row=i + 1, column=15).value = CL_ED_IZM_RAZM
        sheet.cell(row=i + 1, column=16).value = elem['shirina']
        sheet.cell(row=i + 1, column=17).value = CL_ED_IZM_RAZM
        sheet.cell(row=i + 1, column=18).value = elem['visota']
        sheet.cell(row=i + 1, column=19).value = CL_ED_IZM_RAZM
        sheet.cell(row=i + 1, column=20).value = elem['volume']
        sheet.cell(row=i + 1, column=21).value = CL_ED_IZM_VOLUME
        sheet.cell(row=i + 1, column=22).value = elem['weight']
        sheet.cell(row=i + 1, column=23).value = CL_ED_IZM_WEIGHT
        sheet.cell(row=i + 1, column=24).value = elem['country']
        sheet.cell(row=i + 1, column=25).value = elem['srok']
        sheet.cell(row=i + 1, column=26).value = elem['rec_price']
        sheet.cell(row=i + 1, column=27).value = elem['zak_price']
        sheet.cell(row=i + 1, column=28).value = elem['stock']
        sheet.cell(row=i + 1, column=29).value = CL_CURRENCY
        sheet.cell(row=i + 1, column=30).value = CL_CURRENCY
        i += 1
    tm = time.strftime('%d.%m.%Y_%H.%M', time.localtime())
    res_file = DIR + CL_PRIKAT_NAME + "_" + tm + ".xlsx"
    wb.save(res_file)
    return res_file


def make_csv(res: List) -> str:
    i = 1
    #tm = time.strftime('%d.%m.%Y_%H.%M', time.localtime())
    #res_file = DIR + CL_PRIKAT_NAME + "_" + tm + ".csv"
    res_file = CL_PRIKAT_NAME + ".csv"
    with open(res_file, "w", newline='', encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        for elem in res:
            line = [CL_OUR_GLN, i, elem['barcode'], None, elem['code'], elem['name'], 1, CL_ED_IZM, elem['group'], elem['brand'], None, None, None, elem['glubina'],
            CL_ED_IZM_RAZM, elem['shirina'], CL_ED_IZM_RAZM, elem['visota'], CL_ED_IZM_RAZM, elem['volume'], CL_ED_IZM_VOLUME, elem['weight'], CL_ED_IZM_WEIGHT,
            elem['country'], elem['srok'], elem['rec_price'], elem['zak_price'], elem['stock'], CL_CURRENCY, CL_CURRENCY]
            writer.writerow(line)
            i += 1
    return res_file


def get_data_from_xls(filename):
    """
    Разовый модуль заполнения полей МС из прайс-листа из cislink
    """
    wb = openpyxl.load_workbook(DIR + filename)
    sheet = wb['Pricat_AllTools_38']

    for i in range(12, 621):
        line = []
        for j in range(1, 30):
            line.append(sheet.cell(row=i, column=j).value)
        id = ms.get_id_by_code(line[3])
        if not id:
            code = str(line[3])
            code = '0' + code
            id = ms.get_id_by_code(code)
            if id:
                ms.modify_product(id, line)
            else:
                code = '0' + code
                id = ms.get_id_by_code(code)
                if id:
                    ms.modify_product(id, line)
        else:
            ms.modify_product(id, line)
        print(id, line)


def main():
    # get_data_from_xls("source.xlsx")
    res = ms.get_all_stock()
    tm = time.strftime('%d.%m.%Y_%H.%M', time.localtime())
    print(tm, "Получено из МС, товаров -", len(res))
    # file_price = make_xlsx(res)
    file_price = make_csv(res)
    # file_price = "Pricat_AllTools_31.08.2021_16.06.csv"
    error = 0
    i = 0
    while i < TRY:
        try:
            ftp = FTP(URL_FTP)
            ftp.login(user=CL_USER, passwd=CL_PASSWD)
            ftp.cwd(CL_PRICAT_FOLDER)
        except:
            error = 1
            i += 1
        else:
            error = 0
            break
    if error:
        print("Error ftp login")
        sys.exit()
    i = 0
    while i < TRY:
        try:
            with open(DIR + file_price, 'rb') as f:
                ftp.storbinary('STOR ' + file_price, f, blocksize=1024)
        except:
            error = 1
            i += 1
        else:
            error = 0
            break
    ftp.quit()
    if error:
        print("Error uploading pricelist on ftp")
        sys.exit()
    print(f"Файл {file_price} загружен на FTP.")
    # os.remove(DIR + file_price)
    sys.exit()


if __name__ == "__main__":
    main()